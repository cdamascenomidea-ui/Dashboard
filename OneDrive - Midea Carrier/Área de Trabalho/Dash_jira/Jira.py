# Atualizado: usa Path/argparse, suporte a locale para fórmulas (en/pt), valida input e protege divisão por zero
# Uso: python Jira.py input.xlsx [output.xlsx] [--locale en|pt]
from pathlib import Path
import argparse
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

FUNC_MAP = {
    "en": {"count": "COUNTA", "countif": "COUNTIF", "avg": "AVERAGE", "if": "IF", "sep": ","},
    "pt": {"count": "CONT.VALORES", "countif": "CONT.SE", "avg": "MÉDIA", "if": "SE", "sep": ";"},
}


def build_dashboard(input_path: Path, output_path: Path, loc: str = "en") -> Path:
    fmt = FUNC_MAP.get(loc[:2], FUNC_MAP["en"])

    if not input_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {input_path}")

    wb = load_workbook(str(input_path))
    data_ws = wb.active
    data_ws.title = "Base_Dados"

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    dash = wb.create_sheet("Dashboard")

    dash["A1"] = "DASHBOARD OPERACIONAL – CHAMADOS"
    dash["A1"].font = Font(bold=True)

    dash["A3"] = "Total de Chamados"
    dash["A4"] = "Chamados Fora do SLA"
    dash["A5"] = "% Fora do SLA"
    dash["A6"] = "Média Dias em Atraso"

    sep = fmt["sep"]
    dash["B3"] = f"={fmt['count']}(Base_Dados!A:A)"
    dash["B4"] = f'={fmt["countif"]}(Base_Dados!G:G{sep}"Fora do SLA")'
    # Protege divisão por zero
    dash["B5"] = f"={fmt['if']}(B3=0{sep}0{sep}B4/B3)"
    dash["B6"] = f"={fmt['avg']}(Base_Dados!F:F)"

    for row in range(3, 7):
        dash[f"A{row}"].font = Font(bold=True)

    dash["B5"].number_format = "0.00%"

    for col in range(1, 3):
        dash.column_dimensions[get_column_letter(col)].width = 40

    # Ensure output dir exists
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Gera dashboard a partir de um arquivo de chamados Jira")
    parser.add_argument("input", type=Path, help="Caminho para o arquivo Excel de entrada")
    parser.add_argument("output", type=Path, nargs="?", default=Path("Dashboard_Chamados_Jira.xlsx"), help="Arquivo Excel de saída")
    parser.add_argument("--locale", choices=["en", "pt"], default="en", help="Locale para fórmulas do Excel (en/pt)")
    args = parser.parse_args()

    try:
        out = build_dashboard(args.input, args.output, args.locale)
        print(f"Dashboard salvo em: {out}")
    except Exception as e:
        print(f"Erro: {e}")
        sys.exit(1)

