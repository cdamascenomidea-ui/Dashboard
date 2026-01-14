from pathlib import Path
from typing import Union

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

FUNC_MAP = {
    "en": {"count": "COUNTA", "countif": "COUNTIF", "avg": "AVERAGE", "if": "IF", "sep": ","},
    "pt": {"count": "CONT.VALORES", "countif": "CONT.SE", "avg": "MÉDIA", "if": "SE", "sep": ";"},
}


def ensure_dir(path: Path) -> None:
    path = Path(path)
    path.mkdir(parents=True, exist_ok=True)


def load_workbook_path(path: Union[str, Path]):
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")
    return load_workbook(str(path))


def save_workbook(wb, output_path: Union[str, Path]):
    output_path = Path(output_path)
    ensure_dir(output_path.parent)
    wb.save(str(output_path))
    return output_path


def fmt_for_locale(loc: str = "en"):
    return FUNC_MAP.get(loc[:2].lower(), FUNC_MAP["en"])


def safe_division_formula(numer_cell: str, denom_cell: str, loc: str = "en") -> str:
    fmt = fmt_for_locale(loc)
    sep = fmt["sep"]
    # IF(denom=0,0,numer/denom)  or SE(denom=0;0;numer/denom)
    return f"={fmt['if']}({denom_cell}=0{sep}0{sep}{numer_cell}/{denom_cell})"


def build_dashboard(input_path: Union[str, Path], output_path: Union[str, Path], loc: str = "en") -> Path:
    fmt = fmt_for_locale(loc)
    sep = fmt["sep"]

    wb = load_workbook_path(input_path)
    data_ws = wb.active
    data_ws.title = "Base_Dados"

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    dash = wb.create_sheet("Dashboard")

    dash["A1"] = "DASHBOARD OPERACIONAL – CHAMADOS"
    dash["A1"].font = Font(bold=True)

    dash["A3"] = "Total de Chamados"
    dash["A4"] = "Chamados Fora do SLA"
    dash["A5"] = "% Fora do SLA"
    dash["A6"] = "Média Dias em Atraso"

    dash["B3"] = f"={fmt['count']}(Base_Dados!A:A)"
    dash["B4"] = f"={fmt['countif']}(Base_Dados!G:G{sep}\"Fora do SLA\")"
    dash["B5"] = safe_division_formula("B4", "B3", loc)
    dash["B6"] = f"={fmt['avg']}(Base_Dados!F:F)"

    for row in range(3, 7):
        dash[f"A{row}"].font = Font(bold=True)

    dash["B5"].number_format = "0.00%"

    for col in range(1, 3):
        dash.column_dimensions[get_column_letter(col)].width = 40

    out = save_workbook(wb, output_path)
    return out
