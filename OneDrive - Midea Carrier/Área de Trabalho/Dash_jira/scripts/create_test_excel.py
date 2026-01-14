from openpyxl import Workbook
from pathlib import Path


def create_test_file(path: Path = Path("test_chamados.xlsx")):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados"
    # Headers
    ws['A1'] = 'Pedido'
    ws['F1'] = 'Dias em Atraso'
    ws['G1'] = 'SLA'

    rows = [
        ('P1', 2, 'Dentro do SLA'),
        ('P2', 0, 'Fora do SLA'),
        ('P3', 1, 'Fora do SLA'),
        ('P4', 3, 'Dentro do SLA'),
        ('P5', 5, 'Fora do SLA'),
    ]

    for i, (pedido, dias, sla) in enumerate(rows, start=2):
        ws[f'A{i}'] = pedido
        ws[f'F{i}'] = dias
        ws[f'G{i}'] = sla

    wb.save(path)
    print(f"Arquivo de teste criado: {path}")


if __name__ == "__main__":
    create_test_file()
